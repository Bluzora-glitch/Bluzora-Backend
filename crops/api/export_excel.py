import io
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rest_framework.decorators import api_view
from rest_framework.response import Response
import openpyxl
from openpyxl.chart import LineChart, Reference
import urllib.parse
from crops.models import Crop, CropVariable, PredictedData

@api_view(['GET'])
def export_price_data_excel(request):
    # รับ query parameter
    vegetable_name = request.GET.get('vegetableName')
    start_date_str = request.GET.get('startDate')
    end_date_str = request.GET.get('endDate')
    
    if not (vegetable_name and start_date_str and end_date_str):
        return Response({"error": "Missing parameters"}, status=400)
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    crop_obj = Crop.objects.filter(crop_name__icontains=vegetable_name.strip()).first()
    if not crop_obj:
        return Response({"error": "Crop not found"}, status=404)
    
    # ดึงข้อมูล historical และ predicted
    historical_qs = CropVariable.objects.filter(
        crop=crop_obj,
        date__gte=start_date,
        date__lte=end_date
    ).values('date', 'min_price', 'max_price', 'average_price')
    
    predicted_qs = PredictedData.objects.filter(
        crop=crop_obj,
        predicted_date__gte=start_date,
        predicted_date__lte=end_date
    ).values('predicted_date', 'predicted_price')
    
    # รวมข้อมูลทั้ง historical และ predicted เป็นรายการเดียวกัน
    combined_rows = []
    
    # เพิ่มข้อมูล historical ทั้งหมด
    for item in historical_qs:
        date_dt = item['date']
        combined_rows.append({
            "crop_name": crop_obj.crop_name,
            "date": date_dt.strftime("%Y-%m-%d"),
            "date_dt": date_dt,
            "min_price": float(item['min_price']),
            "max_price": float(item['max_price']),
            "average_price": float(item['average_price']),
            "predicted_price": ""  # ไม่มีข้อมูล predicted ในแถว historical
        })
    
    # สร้าง set ของวันที่ที่มีข้อมูล historical (ในรูปแบบสตริง)
    historical_dates = {row["date"] for row in combined_rows}
    
    # เพิ่มข้อมูล predicted เฉพาะวันที่ที่ไม่มีข้อมูล historical
    for item in predicted_qs:
        date_dt = item['predicted_date']
        date_str = date_dt.strftime("%Y-%m-%d")
        if date_str in historical_dates:
            continue  # ข้ามถ้ามีข้อมูล historical อยู่แล้วในวันนั้น
        combined_rows.append({
            "crop_name": crop_obj.crop_name,
            "date": date_str,
            "date_dt": date_dt,
            "min_price": "",
            "max_price": "",
            "average_price": "",
            "predicted_price": float(item['predicted_price'])
        })
    
    # เรียงลำดับ combined_rows โดยใช้ค่า date_dt จากเก่าไปใหม่
    combined_rows.sort(key=lambda row: row["date_dt"])
    
    # สร้าง workbook ด้วย openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Data"

    # เพิ่ม header row
    headers = ["Crop Name", "Date", "Min Price", "Max Price", "Average Price", "Predicted Price"]
    ws.append(headers)

    # เขียนข้อมูลจาก combined_rows ลงใน worksheet
    for row in combined_rows:
        ws.append([
            row["crop_name"],
            row["date"],
            row["min_price"],
            row["max_price"],
            row["average_price"],
            row["predicted_price"],
        ])
    
    # สร้างกราฟ LineChart สำหรับแสดงเฉพาะ Average Price กับ Date
    chart = LineChart()
    chart.title = "Price Forecast"
    chart.y_axis.title = "Price"
    chart.x_axis.title = "Date"

    # กำหนดข้อมูลสำหรับกราฟ: 
    # ใช้คอลัมน์ 5 ("Average Price") สำหรับข้อมูล y-values โดยข้าม header row (min_row=2)
    data = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=False)
    
    # กำหนด categories (แกน x) โดยใช้คอลัมน์ 2 ("Date")
    categories = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    chart.set_categories(categories)
    
    # วางกราฟใน worksheet ที่ตำแหน่ง "H2"
    ws.add_chart(chart, "H2")

    # สร้างไฟล์ excel ใน memory stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"PriceData_{crop_obj.crop_name}_{start_date_str}_to_{end_date_str}.xlsx"
    response = HttpResponse(
        stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = "attachment; filename*=UTF-8''" + urllib.parse.quote(filename)
    return response
